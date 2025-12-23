"""
:Date        : 2025-12-08 10:56:48
:LastEditTime: 2025-12-23 23:09:37
:Description : 
"""
import os
import json
import tempfile
from pathlib import Path

import cv2
import numpy as np
from skimage.io import imsave
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from PyQt5.QtCore import QFileInfo
from ultralytics import YOLO
from aicsimageio.aics_image import AICSImage

from core.measurement_3d import Measurer3D
from .image_processor import merge_channels, stretch_contrast

font = Font(name='Microsoft Yahei', size=11)
alignment = Alignment(horizontal='right', vertical='center')

def parse_json(path: str):
    """
    json to python dict
    """
    if os.path.exists(path):
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    return None

def get_cilia_channel_index(aics_image: AICSImage, channel_params: dict):
    """
    extract channel type from channel_params to find the index for cilia channel
    """
    channel_count = len(aics_image.channel_names)
    for i in range(channel_count):
        if i in channel_params:
            params = channel_params[i]
            if params["type"] == 2:
                return i
    return None

def generate_and_save_projection(aics_image: AICSImage, channel_params, color_definition, save_path):
    """
    Save the MIP image as a named temp file
    """
    raw_data = aics_image.get_image_data("TCZYX")
    max_projection = np.max(raw_data[0, :, :, :, :], axis=1)
    channel_count = len(aics_image.channel_names)
    stretched_channels = []
    channel_colors = []

    for i in range(channel_count):
        if i in channel_params:
            params = channel_params[i]
            channel_data = max_projection[i]

            stretched = stretch_contrast(
                channel_data,
                params['lower'],
                params['upper']
            )
            stretched_channels.append(stretched)
            if params["type"] == 1:
                nuclei_color = [item / 255 for item in color_definition["Nuclei"]]
                channel_colors.append(tuple(nuclei_color))
            elif params["type"] == 2:
                cilia_color = [item / 255 for item in color_definition["Cilia"]]
                channel_colors.append(tuple(cilia_color))
            else:
                channel_colors.append((0.8, 0.8, 0.8))

    rgb_image = merge_channels(stretched_channels, channel_colors)
    imsave(save_path, (rgb_image * 255).astype(np.uint8))

def parse_channel_params(channel_params: dict):
    """
    parse channel params and output to result
    """
    result = ''
    for _, value in channel_params.items():
        value: dict
        for key, setting in value.items():
            if key == "name":
                result += setting
                result += ":\n"
            elif key == "type":
                result += "    type: "
                result += "Nuclei" if setting == 1 else "Cilia"
            else:
                result += "    "
                result += key
                result += ":"
                result += str(setting)
                result += "\n"
        result += "\n"
    return result

def cilia_analysis(
        image_path: str,
        image_snapshot,
        channel_params,
        output_dir
    ):
    """
    Use YOLOv8 and something fancy based on opencv to do the trick.
    """
    model_conf : dict = parse_json("config.json")
    model_name = model_conf["Model"]
    color_definition = model_conf["Model Color Definition"]

    # For exporting results to png file.
    output_img = cv2.cvtColor(image_snapshot, cv2.COLOR_RGB2BGR)
    aics_image = AICSImage(image_path)
    raw_data = aics_image.get_image_data("TCZYX")
    cilia_channel_index = get_cilia_channel_index(aics_image, channel_params)
    
    # Name for output. Including sub directory, png file, and spread sheet
    file_info = QFileInfo(image_path)
    base_name = file_info.baseName()
    tag_output_path = os.path.join(output_dir, f"{base_name}.xlsx")
    img_output_path = os.path.join(output_dir, f"{base_name}.png")

    model = YOLO(f"models/{model_name}")
    measurer3d = Measurer3D(
        aics_image.physical_pixel_sizes.X,
        aics_image.physical_pixel_sizes.Z,
        model_conf.get("Brightness Decay Limit", 0)
    )
    # For exporting results to spread sheet.
    tag_list = []

    # Generate a tempfile for YOLOv8 to analyse.
    tmp_image_path = tempfile.NamedTemporaryFile(suffix='.png', delete=False).name
    generate_and_save_projection(aics_image, channel_params, color_definition, tmp_image_path)

    # YOLOv8 
    results = model.predict(
        source=tmp_image_path,
        conf=model_conf["Minimal Confidence"],
        iou=model_conf["Intersection over Union Threshold"],
        save=False,
        project=output_dir,
        name="",
        exist_ok=True,
        show_labels=False,
        show_conf=False,
        line_width=2
    )
    for result in results:
        boxes = result.boxes
        # If a cell is at the edge of the image, incomplete, and without cilia, drop it.
        dropcase_count = 0
        for index_id, box in enumerate(boxes, start=1):
            box_list = box.xyxy[0].tolist()
            for i, coordinate in enumerate(box_list):
                box_list[i] = int(coordinate)
            x1, y1, x2, y2 = box_list
            classid = int(box.cls)
            height, width, _ = output_img.shape
            cilia_length_2d = 0
            cilia_length_3d = 0
            if classid:
                # ROI, but only cilia channel.
                cilia_roi = raw_data[0, cilia_channel_index, :, y1:y2, x1:x2]
                cleaned_roi = stretch_contrast(
                    cilia_roi,
                    channel_params[cilia_channel_index]['lower'],
                    channel_params[cilia_channel_index]['upper']
                )
                measure_result = measurer3d.measure_from_3d_roi(cleaned_roi)
                cilia_length_2d = measure_result["length_2d_um"]
                cilia_length_3d = measure_result["length_3d_um"]
                img_id = str(index_id - dropcase_count).zfill(3)
                if not os.path.exists(f"{output_dir}/{base_name}/"):
                    os.mkdir(f"{output_dir}/{base_name}/")
                if cilia_length_3d:
                    measurer3d.visualize_3d_measurement(
                        measure_result,
                        f"{output_dir}/{base_name}/{base_name}_{img_id}_3d.png"
                    )
                else:
                    classid = 0
            # If a cell is at the edge of the image, incomplete, and without cilia, drop it.
            if any(i in box_list for i in (0, height, width)):
                if not int(cilia_length_3d):
                    dropcase_count += 1
                    continue
            # PNG file label.
            rectangle_color = (0, 255, 0) if classid else (0, 0, 255)
            cv2.rectangle(output_img, (x1, y1), (x2, y2), rectangle_color, 2)
            label = str(index_id - dropcase_count)
            label_offset = x1 if classid else x2 - 20
            label_color = (255, 255, 255) if classid else (0, 255, 255)
            if y1 - 28 > 0:
                cv2.putText(output_img, label, (label_offset, y1 - 8), cv2.FONT_HERSHEY_SIMPLEX, 0.8,label_color, 2)
            else:
                cv2.putText(output_img, label, (label_offset, y1 + 20), cv2.FONT_HERSHEY_SIMPLEX, 0.8, label_color, 2)
            classname = "Y" if classid else "N"
            tag_list.append([index_id - dropcase_count, classname, str(box_list), cilia_length_2d, cilia_length_3d])
        cv2.imwrite(img_output_path, output_img)
    
        template_excel = load_workbook("templates/cilia.xlsx")
        template_sheet = template_excel.active
        template_sheet['H2'] = parse_channel_params(channel_params)
        for row_idx, row_data in enumerate(tag_list, start=5):
            for col_idx, cell_value in enumerate(row_data, start=1):
                cell = template_sheet.cell(row=row_idx, column=col_idx, value=cell_value)
                cell.font = font
                cell.alignment = alignment
        if os.path.exists(tag_output_path):
            os.remove(tag_output_path)
        template_excel.save(tag_output_path)
        template_excel.close()
    Path(tmp_image_path).unlink(missing_ok=True)
